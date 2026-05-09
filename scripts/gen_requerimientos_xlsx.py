from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ==========================================
# HOJA 1: Requerimientos
# ==========================================
ws = wb.active
ws.title = "Requerimientos"

headers = ["Categoria", "Requerimiento", "Descripcion", "Responsable", "Criticidad", "Estado", "Observaciones"]

data = [
    # 1. Dominio web
    ("1. Dominio web", "Subdominio para la plataforma", "Configurar subdominio ej. plataforma.mso.cl o tpt.mso.cl", "TI Cliente", "Alta", "Pendiente", "Requerido para acceso publico"),
    ("1. Dominio web", "Registro DNS CNAME", "CNAME apuntando a platoniaaa.github.io (hosting actual)", "TI Cliente", "Alta", "Pendiente", "Propagacion 15min a 24h"),
    ("1. Dominio web", "Certificado SSL (HTTPS)", "Automatico via GitHub Pages una vez configurado el CNAME", "MSO", "Alta", "Pendiente", "No requiere accion adicional"),

    # 2. Correos Resend
    ("2. Correos transaccionales", "Verificar dominio en Resend", "Verificar mso.cl o procesos360.cl en panel de Resend", "MSO + TI Cliente", "Alta", "Pendiente", "Definir dominio a usar"),
    ("2. Correos transaccionales", "Registro SPF (TXT)", "1 registro SPF en DNS del dominio elegido", "TI Cliente", "Alta", "Pendiente", "Lo entrega Resend al agregar dominio"),
    ("2. Correos transaccionales", "Registros DKIM (CNAME)", "2 registros DKIM en DNS del dominio elegido", "TI Cliente", "Alta", "Pendiente", "Los entrega Resend"),
    ("2. Correos transaccionales", "Registro DMARC (TXT)", "Recomendado para mejor entregabilidad", "TI Cliente", "Media", "Pendiente", "Opcional"),
    ("2. Correos transaccionales", "Definir email remitente", "Sugerido: no-reply@mso.cl con display name 'MSO Plataforma'", "MSO", "Alta", "Pendiente", "Aplica a bienvenida, reset password, encuestas, recordatorios, confirmaciones, coevaluacion"),

    # 3. Allowlist antispam
    ("3. Allowlist antispam", "Permitir remitente", "no-reply@mso.cl (post verificacion)", "TI Cliente", "Alta", "Pendiente", "Evita cuarentena de correos"),
    ("3. Allowlist antispam", "Permitir enlaces de la plataforma", "Dominio plataforma.mso.cl o el que se defina", "TI Cliente", "Media", "Pendiente", ""),

    # 4. Backend Supabase
    ("4. Backend (Supabase)", "Plan Supabase Pro", "USD 25/mes - 8GB DB, 100GB Storage, 250GB egress, backups 7 dias", "MSO", "Alta", "Pendiente", "Actualmente en plan Free"),
    ("4. Backend (Supabase)", "Definir region", "us-east-1 actual o migrar a sa-east-1 (Sao Paulo) para menor latencia", "MSO", "Media", "Pendiente", "Evaluar costo de migracion"),

    # 5. IA Groq
    ("5. Servicio de IA", "Cuenta Groq API", "Modelo llama-3.3-70b-versatile, actualmente en plan gratuito", "MSO", "Media", "Configurado", "Evaluar upgrade segun volumen de informes"),

    # 6. Datos iniciales
    ("6. Datos iniciales", "Listado de usuarios", "Nombre, Email, Cargo, Rol (lider/colaborador), Relacion lider-colaborador", "Cliente (Sodexo)", "Alta", "Pendiente", "Formato Excel, ya hay plantilla"),
    ("6. Datos iniciales", "Programa a cargar", "Nombre, fechas, descripcion", "Cliente", "Alta", "Pendiente", ""),
    ("6. Datos iniciales", "Carta Gantt del programa", "Excel con actividades y fechas", "Cliente", "Alta", "Pendiente", "Plantilla disponible"),
    ("6. Datos iniciales", "Competencias y conductas", "Excel con competencias a evaluar", "Cliente + MSO", "Alta", "Pendiente", ""),
    ("6. Datos iniciales", "Logo del cliente", "Para branding de la plataforma", "Cliente", "Baja", "Pendiente", "Opcional"),

    # 7. Seguridad y compliance
    ("7. Seguridad", "Migrar contrasenas a bcrypt", "Migrar a Supabase Auth con hash bcrypt (actualmente en texto plano)", "MSO", "Critica", "Pendiente", "Obligatorio para produccion"),
    ("7. Seguridad", "HTTPS obligatorio", "Automatico con dominio + GitHub Pages", "MSO", "Alta", "Configurado", ""),
    ("7. Seguridad", "Revisar RLS Supabase", "Row Level Security en todas las tablas antes de go-live", "MSO", "Critica", "Pendiente", "Auditoria completa"),
    ("7. Seguridad", "Politica de privacidad", "Redactar y publicar en la plataforma", "MSO + Cliente", "Alta", "Pendiente", "Ley 19.628 Chile"),
    ("7. Seguridad", "Terminos de uso", "Redactar y publicar en la plataforma", "MSO + Cliente", "Alta", "Pendiente", ""),
    ("7. Seguridad", "Consentimiento explicito", "Checkbox al registrarse / primer login", "MSO", "Alta", "Pendiente", "Ley 19.628"),

    # 8. Navegadores
    ("8. Usuario final", "Navegadores soportados", "Chrome, Edge, Firefox, Safari (ultimas 2 versiones)", "Cliente", "Media", "Informativo", ""),
    ("8. Usuario final", "JavaScript habilitado", "Requerido para funcionamiento de la plataforma", "Cliente", "Alta", "Informativo", "Estandar en empresas"),
    ("8. Usuario final", "Conexion a internet estable", "No requiere ancho de banda elevado", "Cliente", "Media", "Informativo", ""),

    # 9. Monitoreo y SLA
    ("9. Monitoreo y SLA", "Dashboard Supabase", "Acceso a logs, uso, errores, queries", "MSO", "Media", "Configurado", ""),
    ("9. Monitoreo y SLA", "Canal de soporte MSO", "Email de soporte y horario definido", "MSO", "Alta", "Pendiente", "Definir email oficial"),
    ("9. Monitoreo y SLA", "SLA de respuesta", "Respuesta en 24h habiles, resolucion por severidad", "MSO + Cliente", "Alta", "Pendiente", "Acordar con cliente"),
    ("9. Monitoreo y SLA", "Ventana de mantenimiento", "Definir horario (ej. domingos 02:00-04:00 CLT)", "MSO + Cliente", "Media", "Pendiente", ""),

    # 10. Responsables
    ("10. Responsables", "Administrador tecnico MSO", "Nombre, email, telefono", "MSO", "Alta", "Pendiente", "Definir contraparte"),
    ("10. Responsables", "Contraparte TI cliente", "Nombre, email, telefono", "Cliente", "Alta", "Pendiente", ""),
    ("10. Responsables", "Acceso a panel DNS", "Quien administra el dominio del cliente", "TI Cliente", "Alta", "Pendiente", ""),
    ("10. Responsables", "Credenciales Supabase", "Solo equipo tecnico MSO", "MSO", "Critica", "Configurado", "Secretos no compartidos"),
]

# Colores por categoria
cat_colors = {
    "1. Dominio web":            "D4E6F1",
    "2. Correos transaccionales":"D1F2EB",
    "3. Allowlist antispam":     "FCF3CF",
    "4. Backend (Supabase)":     "E8DAEF",
    "5. Servicio de IA":         "FADBD8",
    "6. Datos iniciales":        "F5CBA7",
    "7. Seguridad":              "F5B7B1",
    "8. Usuario final":          "AED6F1",
    "9. Monitoreo y SLA":        "ABEBC6",
    "10. Responsables":          "D6DBDF",
}

# Estilos base
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill("solid", fgColor="4A148C")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_align = Alignment(vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Headers
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = border

# Datos
crit_colors = {
    "Critica": "C0392B",
    "Alta":    "E67E22",
    "Media":   "F1C40F",
    "Baja":    "58D68D",
}
estado_colors = {
    "Pendiente":    "F8C471",
    "Configurado":  "82E0AA",
    "Informativo":  "AED6F1",
    "En progreso":  "85C1E9",
}

for i, row in enumerate(data, 2):
    for col, val in enumerate(row, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.alignment = cell_align
        c.border = border
        # Color de fondo por categoria
        if col == 1:
            c.fill = PatternFill("solid", fgColor=cat_colors.get(row[0], "EEEEEE"))
            c.font = Font(bold=True, size=10)
        # Color para criticidad
        if col == 5 and val in crit_colors:
            c.fill = PatternFill("solid", fgColor=crit_colors[val])
            c.font = Font(bold=True, color="FFFFFF" if val in ("Critica","Alta") else "000000", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
        # Color para estado
        if col == 6 and val in estado_colors:
            c.fill = PatternFill("solid", fgColor=estado_colors[val])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(bold=True, size=10)

# Anchos de columna
widths = [24, 38, 55, 22, 14, 16, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 32
ws.freeze_panes = "A2"

# ==========================================
# HOJA 2: Checklist Go-Live
# ==========================================
ws2 = wb.create_sheet("Checklist Go-Live")

checklist = [
    ("Dominio plataforma.mso.cl configurado y apuntando al hosting", "Alta", ""),
    ("Dominio de correo verificado en Resend", "Alta", ""),
    ("Registros DNS (SPF, DKIM, DMARC) validados", "Alta", ""),
    ("Email remitente funcionando en pruebas de envio", "Alta", ""),
    ("Plan Supabase Pro activado", "Alta", ""),
    ("Backups automaticos verificados", "Alta", ""),
    ("Migracion de contrasenas a bcrypt completada", "Critica", ""),
    ("RLS revisado en todas las tablas", "Critica", ""),
    ("Politica de privacidad publicada", "Alta", ""),
    ("Terminos de uso publicados", "Alta", ""),
    ("Consentimiento de uso de datos implementado", "Alta", ""),
    ("Usuarios iniciales cargados (lideres y colaboradores)", "Alta", ""),
    ("Programas, competencias y Gantt cargados", "Alta", ""),
    ("Allowlist de correo aplicado en el cliente", "Alta", ""),
    ("Pruebas end-to-end del flujo completo", "Alta", ""),
    ("Capacitacion al administrador del cliente", "Media", ""),
    ("Canal de soporte comunicado al cliente", "Alta", ""),
    ("Documentacion de usuario entregada", "Media", ""),
]

ws2.cell(row=1, column=1, value="Item").font = header_font
ws2.cell(row=1, column=2, value="Criticidad").font = header_font
ws2.cell(row=1, column=3, value="Completado").font = header_font
ws2.cell(row=1, column=4, value="Observaciones").font = header_font
for col in range(1, 5):
    c = ws2.cell(row=1, column=col)
    c.fill = header_fill
    c.alignment = header_align
    c.border = border

for i, (item, crit, obs) in enumerate(checklist, 2):
    c1 = ws2.cell(row=i, column=1, value=item); c1.alignment = cell_align; c1.border = border
    c2 = ws2.cell(row=i, column=2, value=crit); c2.alignment = Alignment(horizontal="center", vertical="center"); c2.border = border
    if crit in crit_colors:
        c2.fill = PatternFill("solid", fgColor=crit_colors[crit])
        c2.font = Font(bold=True, color="FFFFFF" if crit in ("Critica","Alta") else "000000")
    c3 = ws2.cell(row=i, column=3, value=""); c3.alignment = Alignment(horizontal="center", vertical="center"); c3.border = border
    c4 = ws2.cell(row=i, column=4, value=obs); c4.alignment = cell_align; c4.border = border

ws2.column_dimensions['A'].width = 65
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 40
ws2.row_dimensions[1].height = 30
ws2.freeze_panes = "A2"

# ==========================================
# HOJA 3: DNS Resend (ejemplo)
# ==========================================
ws3 = wb.create_sheet("DNS Resend")

ws3.cell(row=1, column=1, value="Registros DNS requeridos para Resend").font = Font(bold=True, size=14, color="4A148C")
ws3.merge_cells("A1:E1")

dns_headers = ["Tipo", "Host / Name", "Valor", "TTL", "Notas"]
for col, h in enumerate(dns_headers, 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border

dns_rows = [
    ("TXT", "@ (o el dominio raiz)", "v=spf1 include:amazonses.com ~all", "3600", "SPF - sobrescribir si ya existe"),
    ("CNAME", "resend._domainkey", "(valor entregado por Resend al agregar el dominio)", "3600", "DKIM 1"),
    ("CNAME", "resend2._domainkey", "(valor entregado por Resend)", "3600", "DKIM 2"),
    ("TXT", "_dmarc", "v=DMARC1; p=none; rua=mailto:admin@mso.cl", "3600", "DMARC - opcional"),
]
for i, row in enumerate(dns_rows, 4):
    for col, val in enumerate(row, 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.alignment = cell_align; c.border = border

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 55
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 40
ws3.row_dimensions[3].height = 28

ws3.cell(row=9, column=1, value="Importante").font = Font(bold=True, size=12, color="C0392B")
ws3.cell(row=10, column=1, value="Los valores exactos de DKIM se obtienen al agregar el dominio en el panel de Resend.").alignment = Alignment(wrap_text=True)
ws3.merge_cells("A10:E10")

# ==========================================
# HOJA 4: Glosario / Tecnico
# ==========================================
ws4 = wb.create_sheet("Glosario Tecnico")
gl = [
    ("Termino", "Definicion"),
    ("SPF", "Sender Policy Framework - Indica que servidores pueden enviar correo en nombre del dominio"),
    ("DKIM", "DomainKeys Identified Mail - Firma digital que verifica la autenticidad del correo"),
    ("DMARC", "Domain-based Message Authentication - Politica que indica que hacer con correos que fallan SPF/DKIM"),
    ("CNAME", "Canonical Name Record - Alias DNS que apunta un dominio a otro"),
    ("TXT", "Text Record - Registro DNS para texto arbitrario (usado por SPF, DMARC)"),
    ("TTL", "Time To Live - Tiempo que los servidores DNS cachean el registro (en segundos)"),
    ("RLS", "Row Level Security - Seguridad a nivel de fila en Supabase/PostgreSQL"),
    ("bcrypt", "Algoritmo de hashing para contrasenas, resistente a ataques de fuerza bruta"),
    ("Supabase", "Backend as a Service basado en PostgreSQL, Storage y Edge Functions"),
    ("Resend", "Servicio de entrega de correos transaccionales por API"),
    ("Groq", "Servicio de inferencia de modelos de IA por API (usa Llama 3.3)"),
    ("GitHub Pages", "Hosting gratuito de sitios estaticos provisto por GitHub"),
    ("Ley 19.628", "Ley chilena de Proteccion de Datos Personales"),
]
for i, row in enumerate(gl, 1):
    for col, val in enumerate(row, 1):
        c = ws4.cell(row=i, column=col, value=val)
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if i == 1:
            c.font = header_font; c.fill = header_fill; c.alignment = header_align
ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 90
ws4.row_dimensions[1].height = 28
ws4.freeze_panes = "A2"

# Save
out = r"c:\Users\the_r\Documents\GitHub\plataforma-mso\docs\Requerimientos_Deploy_Productivo_TPT.xlsx"
wb.save(out)
print("OK:", out)
